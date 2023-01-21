from pprint import pprint

from docxtpl import DocxTemplate
from openpyxl import load_workbook

import datetime

from odf.opendocument import OpenDocumentText, load
from odf.style import ParagraphProperties, Style, TableColumnProperties, TextProperties
from odf.table import Table, TableColumn, TableRow, TableCell
from odf.text import P

from DataBase import DataBase


class XlsxImport:
    def import_into_list(self, file_name):
        wb2 = load_workbook(file_name)
        sheet = wb2.active
        rows = sheet.max_row
        # cols = sheet.max_column
        items = []
        for i in range(7, rows + 1):
            items_vals = []

            for j in [7,8,13,14]:
                u = sheet.cell(row=i, column=j).value
                if (j == 13 and str(u) == '-'):
                    items_vals.append(sheet.cell(row=i, column=15).value)
                    continue
                if (j == 14 and str(u) == '-'):
                    items_vals.append(sheet.cell(row=i, column=16).value)
                    continue
                items_vals.append(sheet.cell(row=i, column=j).value)

            items.append(items_vals)
        items.pop(-1)
        for i in items:
            i.append(round(i[3]/i[2], 2))
        return items

class ODTExport:

    def form_odt(self, file_name, **kwargs):
        textdoc = OpenDocumentText()

        tablecontents = Style(name="Table Contents", family="paragraph")
        tablecontents.addElement(ParagraphProperties(numberlines="false", linenumber="0"))
        textdoc.styles.addElement(tablecontents)

        widthshort = Style(name="Wshort", family="table-column")
        widthshort.addElement(TableColumnProperties(columnwidth="0.5cm"))
        textdoc.automaticstyles.addElement(widthshort)

        widthmed = Style(name="Wmed", family="table-column")
        widthmed.addElement(TableColumnProperties(columnwidth="2.0cm"))
        textdoc.automaticstyles.addElement(widthmed)

        widthwide = Style(name="Wwide", family="table-column")
        widthwide.addElement(TableColumnProperties(columnwidth="5.5cm"))
        textdoc.automaticstyles.addElement(widthwide)

        p = P(text=('Отчёт по затраченным средствам на пациента'))
        textdoc.text.addElement(p)

        p = P(text=('ФИО пациента: {}'.format(kwargs['patient_fio'])))
        textdoc.text.addElement(p)

        p = P(text=('История болезни: {}'.format(kwargs['story'])))
        textdoc.text.addElement(p)

        p = P(text=('Период лечения: c {} по {}'.format(kwargs['from'], kwargs['to'])))
        textdoc.text.addElement(p)

        p = P(text=('Статус пациента: {}'.format(kwargs['status'])))
        textdoc.text.addElement(p)

        p = P(text=('Сумма затраченых средств: {} руб.'.format(kwargs['total_sum'])))
        textdoc.text.addElement(p)

        p = P(text=('Дата подготовки отчёта {:%d.%m.%Y}'.format(datetime.date.today())))
        textdoc.text.addElement(p)

        table = Table()
        table.addElement(TableColumn(numbercolumnsrepeated=1, stylename=widthshort))
        table.addElement(TableColumn(numbercolumnsrepeated=1, stylename=widthwide))
        table.addElement(TableColumn(numbercolumnsrepeated=1, stylename=widthmed))
        table.addElement(TableColumn(numbercolumnsrepeated=1, stylename=widthmed))
        table.addElement(TableColumn(numbercolumnsrepeated=1, stylename=widthmed))
        table.addElement(TableColumn(numbercolumnsrepeated=1, stylename=widthmed))

        col_names = ['№ п/п', 'Наименование', 'Ед. измерения', 'Кол-во', 'Стоимость за еденицу', 'Сумма за позицию']

        tr = TableRow()
        table.addElement(tr)

        for val in col_names:
            tc = TableCell()
            tr.addElement(tc)
            p = P(stylename=tablecontents, text=val)
            tc.addElement(p)

        for pp, lst in enumerate(kwargs['positions']):
            tr = TableRow()
            table.addElement(tr)

            list_of_cols = [str(pp + 1), lst[6], lst[8], lst[3], lst[7], lst[4]]

            for i in list_of_cols:
                tc = TableCell()
                tr.addElement(tc)
                p = P(stylename=tablecontents, text=i)
                tc.addElement(p)

        textdoc.text.addElement(table)

        textdoc.save(file_name)

    def form_odt_for_sum_of(self, data):
        textdoc = OpenDocumentText()

        tablecontents = Style(name="Table Contents", family="paragraph")
        tablecontents.addElement(ParagraphProperties(numberlines="false", linenumber="0"))
        textdoc.styles.addElement(tablecontents)

        widthshort = Style(name="Wshort", family="table-column")
        widthshort.addElement(TableColumnProperties(columnwidth="0.5cm"))
        textdoc.automaticstyles.addElement(widthshort)

        widthmed = Style(name="Wmed", family="table-column")
        widthmed.addElement(TableColumnProperties(columnwidth="2.0cm"))
        textdoc.automaticstyles.addElement(widthmed)

        widthwide = Style(name="Wwide", family="table-column")
        widthwide.addElement(TableColumnProperties(columnwidth="5.5cm"))
        textdoc.automaticstyles.addElement(widthwide)

        table = Table()
        table.addElement(TableColumn(numbercolumnsrepeated=1, stylename=widthwide))
        table.addElement(TableColumn(numbercolumnsrepeated=1, stylename=widthmed))
        table.addElement(TableColumn(numbercolumnsrepeated=1, stylename=widthshort))
        table.addElement(TableColumn(numbercolumnsrepeated=1, stylename=widthmed))
        table.addElement(TableColumn(numbercolumnsrepeated=1, stylename=widthmed))


        col_names = ['Наименование', 'Ед. измерения', 'Кол-во', 'Стоимость за еденицу', 'Сумма за позицию']

        tr = TableRow()
        table.addElement(tr)

        for val in col_names:
            tc = TableCell()
            tr.addElement(tc)
            p = P(stylename=tablecontents, text=val)
            tc.addElement(p)

        for lst in data:
            tr = TableRow()
            table.addElement(tr)

            list_of_cols = [lst[0], lst[1], lst[2], lst[3], lst[4]]

            for i in list_of_cols:
                tc = TableCell()
                tr.addElement(tc)
                p = P(stylename=tablecontents, text=i)
                tc.addElement(p)

        textdoc.text.addElement(table)

        textdoc.save('itog.odt')


def export(context):
    tpl = DocxTemplate('Appendix_template.docx')
    tpl.render(context)
    tpl.save('dynamic_table.docx')






if __name__ == "__main__":
    odt = ODTExport()
